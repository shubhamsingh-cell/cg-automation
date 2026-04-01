"""CG Automation -- FastAPI backend for Craigslist ad campaign posting optimiser."""

from __future__ import annotations

import base64
import io
import json
import logging
import math
import os
import ssl
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
from collections import OrderedDict
from datetime import datetime, timedelta
from typing import Any, Optional

import numpy as np

import anthropic
import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

import engine
import supabase_store

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
logger = logging.getLogger("cg-automation")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MAX_STORED_JOBS: int = 50

# Excel styling colours
CLR_HEADER_BG = "0F2037"
CLR_SUBHEADER = "2E74B5"
CLR_POS_NR = "1E8449"
CLR_NEG_NR = "C0392B"
CLR_TIER1_BG = "D5F5E3"
CLR_TIER2_BG = "D6EAF8"
CLR_TIER3_BG = "FDEBD0"
CLR_TIER4_BG = "EAECEE"
CLR_LOC_MULT = "D6EAF8"
CLR_CAT_MULT = "FEF9E7"

TIER_BG_MAP: dict[int, str] = {
    1: CLR_TIER1_BG,
    2: CLR_TIER2_BG,
    3: CLR_TIER3_BG,
    4: CLR_TIER4_BG,
}

# ---------------------------------------------------------------------------
# FastAPI app
# ---------------------------------------------------------------------------
app = FastAPI(title="CG Automation API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# In-memory job store (LRU, max 50)
# ---------------------------------------------------------------------------


class LRUDict(OrderedDict):
    """OrderedDict that evicts oldest entries when max_size is exceeded."""

    def __init__(self, max_size: int = MAX_STORED_JOBS) -> None:
        super().__init__()
        self.max_size = max_size

    def __setitem__(self, key: str, value: Any) -> None:
        if key in self:
            self.move_to_end(key)
        super().__setitem__(key, value)
        while len(self) > self.max_size:
            self.popitem(last=False)


job_store: LRUDict = LRUDict()

# ---------------------------------------------------------------------------
# Anthropic client
# ---------------------------------------------------------------------------
ANTHROPIC_API_KEY: str = os.environ.get("ANTHROPIC_API_KEY") or ""


def _get_anthropic_client() -> anthropic.Anthropic:
    """Return an Anthropic client; raises if key is missing."""
    if not ANTHROPIC_API_KEY:
        raise HTTPException(
            status_code=500, detail="ANTHROPIC_API_KEY not configured"
        )
    return anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)


# ---------------------------------------------------------------------------
# Slack notifications (optional, non-blocking)
# ---------------------------------------------------------------------------
SLACK_WEBHOOK_URL: str = os.environ.get("SLACK_WEBHOOK_URL") or ""
DEPLOYED_URL: str = os.environ.get("DEPLOYED_URL") or ""


def _send_slack_webhook(payload: dict[str, Any]) -> None:
    """POST a JSON payload to the Slack webhook URL.

    Runs synchronously -- intended to be called from a daemon thread.
    Logs warnings on failure but never raises.
    """
    if not SLACK_WEBHOOK_URL:
        return
    try:
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            SLACK_WEBHOOK_URL,
            data=data,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            if resp.status != 200:
                logger.warning("Slack webhook returned status %d", resp.status)
    except Exception:
        logger.warning("Slack notification failed", exc_info=True)


def _build_analysis_slack_message(result: dict[str, Any], job_id: str) -> dict[str, Any]:
    """Build a Slack message payload summarising the analysis result.

    Args:
        result: Sanitised analysis result dict.
        job_id: UUID of the stored job.

    Returns:
        Slack webhook payload dict with text and blocks.
    """
    dap: list[dict[str, Any]] = result.get("daily_action_plan", [])
    location_count: int = len(dap)

    # Top 5 by Est Lifetime NR
    sorted_dap = sorted(
        dap,
        key=lambda r: float(r.get("est_lifetime_nr", r.get("Est_Lifetime_NR", 0)) or 0),
        reverse=True,
    )
    top5 = sorted_dap[:5]

    top5_lines: list[str] = []
    for i, rec in enumerate(top5, 1):
        tier = rec.get("tier", rec.get("Tier", ""))
        loc = rec.get("location", rec.get("Location", ""))
        title = rec.get("recommended_title", rec.get("Best_Title", ""))
        est_nr = float(rec.get("est_lifetime_nr", rec.get("Est_Lifetime_NR", 0)) or 0)
        top5_lines.append(f"{i}. T{tier} | {loc} | {title} | ${est_nr:,.2f}")

    # Totals from scorecard if available, otherwise sum from DAP
    scorecard: dict[str, Any] = result.get("scorecard", {})
    total_spend = scorecard.get("total_spend", scorecard.get("total_cost"))
    total_nr = scorecard.get("total_nr", scorecard.get("total_lifetime_nr"))

    # Fallback: sum from DAP
    if total_spend is None:
        total_spend = sum(
            float(r.get("d1_cost", r.get("D1_Cost", 0)) or 0) for r in dap
        )
    if total_nr is None:
        total_nr = sum(
            float(r.get("est_lifetime_nr", r.get("Est_Lifetime_NR", 0)) or 0)
            for r in dap
        )

    top5_text = "\n".join(top5_lines) if top5_lines else "(none)"

    download_line = ""
    if DEPLOYED_URL:
        download_line = f"\n<{DEPLOYED_URL}/api/download/{job_id}|Download Excel>"

    text = (
        f"*CG Automation: Daily Action Plan Ready*\n"
        f"Total locations to post: *{location_count}*\n\n"
        f"*Top 5 by Est Lifetime NR:*\n{top5_text}\n\n"
        f"Total Est Spend: *${float(total_spend):,.2f}*\n"
        f"Total Est NR: *${float(total_nr):,.2f}*"
        f"{download_line}"
    )

    return {"text": text}


def _notify_slack_analysis(result: dict[str, Any], job_id: str) -> None:
    """Send Slack notification for a completed analysis in a background thread.

    Args:
        result: Sanitised analysis result dict.
        job_id: UUID of the stored job.
    """
    if not SLACK_WEBHOOK_URL:
        return
    payload = _build_analysis_slack_message(result, job_id)
    thread = threading.Thread(target=_send_slack_webhook, args=(payload,), daemon=True)
    thread.start()


# ---------------------------------------------------------------------------
# Google Sheets export (service account via GOOGLE_SHEETS_CREDENTIALS_B64)
# ---------------------------------------------------------------------------
GOOGLE_SHEETS_CREDENTIALS_B64: str = (
    os.environ.get("GOOGLE_SHEETS_CREDENTIALS_B64") or ""
)
_GSHEETS_SCOPES: list[str] = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]
_GSHEETS_TOKEN_URI: str = "https://oauth2.googleapis.com/token"
_GSHEETS_BASE: str = "https://sheets.googleapis.com/v4/spreadsheets"
_GDRIVE_BASE: str = "https://www.googleapis.com/drive/v3/files"
_gsheets_token_cache: dict[str, Any] = {"token": None, "expires_at": 0.0}


def _load_gsheets_credentials() -> Optional[dict[str, str]]:
    """Load Google service account credentials from base64 env var."""
    if not GOOGLE_SHEETS_CREDENTIALS_B64:
        return None
    required_fields = ("client_email", "private_key", "token_uri")
    try:
        decoded = base64.b64decode(GOOGLE_SHEETS_CREDENTIALS_B64)
        creds: dict[str, str] = json.loads(decoded)
        for field in required_fields:
            if field not in creds:
                logger.error("Google SA JSON missing field: %s", field)
                return None
        return creds
    except Exception as exc:
        logger.error("Failed to decode GOOGLE_SHEETS_CREDENTIALS_B64: %s", exc, exc_info=True)
        return None


def _build_gsheets_jwt(creds: dict[str, str]) -> str:
    """Build a signed RS256 JWT for Google OAuth2 token exchange."""
    from cryptography.hazmat.primitives import hashes, serialization
    from cryptography.hazmat.primitives.asymmetric import padding as rsa_padding

    now = int(time.time())
    header = {"alg": "RS256", "typ": "JWT"}
    clms = {
        "iss": creds["client_email"],
        "scope": " ".join(_GSHEETS_SCOPES),
        "aud": creds.get("token_uri") or _GSHEETS_TOKEN_URI,
        "iat": now, "exp": now + 3600,
    }

    def _b64url(data: bytes) -> str:
        return base64.urlsafe_b64encode(data).rstrip(b"=").decode("ascii")

    hdr_b64 = _b64url(json.dumps(header, separators=(",", ":")).encode())
    clm_b64 = _b64url(json.dumps(clms, separators=(",", ":")).encode())
    si = f"{hdr_b64}.{clm_b64}".encode("ascii")
    try:
        key = serialization.load_pem_private_key(creds["private_key"].encode("utf-8"), password=None)
        sig = key.sign(si, rsa_padding.PKCS1v15(), hashes.SHA256())  # type: ignore[union-attr]
        return f"{hdr_b64}.{clm_b64}.{_b64url(sig)}"
    except Exception as exc:
        raise RuntimeError(f"JWT RS256 signing failed: {exc}") from exc


def _get_gsheets_access_token() -> Optional[str]:
    """Obtain a Google OAuth2 access token. Caches until 5 min before expiry."""
    now = time.time()
    if _gsheets_token_cache["token"] and _gsheets_token_cache["expires_at"] > now + 300:
        return _gsheets_token_cache["token"]
    creds = _load_gsheets_credentials()
    if not creds:
        return None
    try:
        jwt_token = _build_gsheets_jwt(creds)
    except RuntimeError as exc:
        logger.error("Google Sheets JWT failed: %s", exc, exc_info=True)
        return None
    tok_payload = urllib.parse.urlencode({
        "grant_type": "urn:ietf:params:oauth:2.0-jwt-bearer", "assertion": jwt_token,
    }).encode("utf-8")
    token_uri = creds.get("token_uri") or _GSHEETS_TOKEN_URI
    req = urllib.request.Request(token_uri, data=tok_payload,
                                 headers={"Content-Type": "application/x-www-form-urlencoded"})
    try:
        ctx = ssl.create_default_context()
        with urllib.request.urlopen(req, context=ctx, timeout=15) as resp:
            td = json.loads(resp.read().decode("utf-8"))
        _gsheets_token_cache["token"] = td["access_token"]
        _gsheets_token_cache["expires_at"] = now + td.get("expires_in", 3600)
        return _gsheets_token_cache["token"]
    except (urllib.error.URLError, json.JSONDecodeError, KeyError) as exc:
        logger.error("Google OAuth2 token exchange failed: %s", exc, exc_info=True)
        return None


def _gsheets_request(
    method: str, url: str, body: Optional[dict] = None, token: Optional[str] = None,
) -> Optional[dict]:
    """Make an authenticated request to the Google Sheets/Drive API."""
    if token is None:
        token = _get_gsheets_access_token()
    if not token:
        return None
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    data = json.dumps(body).encode("utf-8") if body else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        ctx = ssl.create_default_context()
        with urllib.request.urlopen(req, context=ctx, timeout=30) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        err_body = ""
        try:
            err_body = exc.read().decode("utf-8")
        except Exception:
            pass
        logger.error("Google API %s %s -> %d: %s", method, url, exc.code, err_body, exc_info=True)
        return None
    except (urllib.error.URLError, json.JSONDecodeError) as exc:
        logger.error("Google API request failed: %s", exc, exc_info=True)
        return None


def _safe_cell(value: Any) -> str:
    """Convert a value to a safe string for Google Sheets cells."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return ""
        return str(value)
    s = str(value)
    if s and s[0] in ("=", "+", "@", "-"):
        return f"'{s}"
    return s


def _job_result_to_sheet_data(result: dict[str, Any]) -> dict[str, list[list[str]]]:
    """Convert analysis result dict into sheet-name -> rows mapping (9 sheets)."""
    sheets: dict[str, list[list[str]]] = {}
    # Sheet 1: Daily Action Plan
    dap_h = ["#", "Location", "Best Title", "Best Category", "Best Day",
             "Today Good?", "Tier", "Trigger", "Est D1 NR", "Est Lifetime NR",
             "Posts/Week", "Multiplier", "Mult Source", "Competing Filtered", "Last Run Date"]
    dap_rows: list[list[str]] = [dap_h]
    cc: dict[str, int] = {}
    for c in result.get("location_conflicts", []):
        lk = str(c.get("Location", c.get("location", ""))).strip().lower()
        cc[lk] = cc.get(lk, 0) + 1
    for i, r in enumerate(result.get("daily_action_plan", []), 1):
        loc = r.get("location", r.get("Location", ""))
        lk = str(loc).strip().lower()
        dap_rows.append([str(i), _safe_cell(loc),
            _safe_cell(r.get("recommended_title", r.get("Best_Title", ""))),
            _safe_cell(r.get("recommended_category", r.get("Best_Category", ""))),
            _safe_cell(r.get("best_day", r.get("Best_Day", ""))),
            _safe_cell(r.get("today_good", r.get("Today_Good", ""))),
            _safe_cell(r.get("tier", r.get("Tier", ""))),
            _safe_cell(r.get("trigger_reason", r.get("Trigger_Reason", ""))),
            _safe_cell(r.get("d1_nr", r.get("D1_NR", r.get("Est_D1_NR", 0)))),
            _safe_cell(r.get("est_lifetime_nr", r.get("Est_Lifetime_NR", 0))),
            _safe_cell(r.get("optimal_posts_per_week", r.get("Optimal_Posts_Per_Week", 1))),
            _safe_cell(r.get("multiplier", r.get("Multiplier_Used", ""))),
            _safe_cell(r.get("mult_source", r.get("Mult_Source", ""))),
            str(cc.get(lk, 0)),
            _safe_cell(r.get("last_run_date", r.get("D1_Date", "")))])
    sheets["Daily Action Plan"] = dap_rows
    # Sheet 2: All Repost Candidates
    rp_h = ["#", "Location", "Title", "Category", "Tier", "Cost", "Profit %",
            "Trigger", "D1 Applies", "D1 NR", "Lifetime NR", "Multiplier",
            "Mult Source", "Best Day", "Today Good?", "Posts/Week"]
    rp_rows: list[list[str]] = [rp_h]
    for i, r in enumerate(result.get("all_repost", []), 1):
        p = r.get("Profit_Pct")
        ps = f"{p:.1f}%" if isinstance(p, (int, float)) else _safe_cell(p)
        rp_rows.append([str(i), _safe_cell(r.get("Location", "")),
            _safe_cell(r.get("Title", "")), _safe_cell(r.get("Category", "")),
            _safe_cell(r.get("Tier", "")), _safe_cell(r.get("D1_Cost", 0)), ps,
            _safe_cell(r.get("Trigger_Reason", "")), _safe_cell(r.get("D1_Applies", 0)),
            _safe_cell(r.get("D1_NR", 0)), _safe_cell(r.get("Est_Lifetime_NR", 0)),
            _safe_cell(r.get("Multiplier_Used", "")), _safe_cell(r.get("Mult_Source", "")),
            _safe_cell(r.get("Best_Day", "")), _safe_cell(r.get("Today_Good", "")),
            _safe_cell(r.get("Optimal_Posts_Per_Week", 1))])
    sheets["All Repost"] = rp_rows
    # Sheet 3: Best Per Location
    bpl_h = ["#", "Location", "Best Title", "Best Category", "Tier",
             "Est Lifetime NR", "Best Day", "Posts/Week"]
    bpl_rows: list[list[str]] = [bpl_h]
    for i, r in enumerate(result.get("best_per_location", []), 1):
        bpl_rows.append([str(i), _safe_cell(r.get("Location", "")),
            _safe_cell(r.get("Title", "")), _safe_cell(r.get("Category", "")),
            _safe_cell(r.get("Tier", "")), _safe_cell(r.get("Est_Lifetime_NR", 0)),
            _safe_cell(r.get("DayOfWeek_Posted", "")), _safe_cell(r.get("Run_Length", ""))])
    sheets["Best Per Location"] = bpl_rows
    # Sheet 4: Location Conflicts
    lc_h = ["Location", "Title", "Category", "Tier", "Est Lifetime NR", "Lost To", "NR Gap"]
    lc_rows: list[list[str]] = [lc_h]
    for r in result.get("location_conflicts", []):
        lc_rows.append([_safe_cell(r.get("Location", "")), _safe_cell(r.get("Title", "")),
            _safe_cell(r.get("Category", "")), _safe_cell(r.get("Tier", "")),
            _safe_cell(r.get("Est_Lifetime_NR", 0)), _safe_cell(r.get("lost_to", "")),
            _safe_cell(r.get("nr_gap", 0))])
    sheets["Location Conflicts"] = lc_rows
    # Sheet 5: Keep Running
    kr_h = ["Location", "Title", "Category", "Total NR", "Profit %", "Trigger"]
    kr_rows: list[list[str]] = [kr_h]
    for r in result.get("keep_running", []):
        p = r.get("Profit_Pct")
        ps = f"{p:.1f}%" if isinstance(p, (int, float)) else _safe_cell(p)
        kr_rows.append([_safe_cell(r.get("Location", "")), _safe_cell(r.get("Title", "")),
            _safe_cell(r.get("Category", "")), _safe_cell(r.get("Total_NR", 0)), ps,
            _safe_cell(r.get("Trigger_Reason", ""))])
    sheets["Keep Running"] = kr_rows
    # Sheet 6: Skip
    sk_h = ["Location", "Title", "Category", "Total NR", "Profit %", "Skip Reason"]
    sk_rows: list[list[str]] = [sk_h]
    for r in result.get("skip", []):
        p = r.get("Profit_Pct")
        ps = f"{p:.1f}%" if isinstance(p, (int, float)) else _safe_cell(p)
        sk_rows.append([_safe_cell(r.get("Location", "")), _safe_cell(r.get("Title", "")),
            _safe_cell(r.get("Category", "")), _safe_cell(r.get("Total_NR", 0)), ps,
            _safe_cell(r.get("Trigger_Reason", ""))])
    sheets["Skip"] = sk_rows
    # Sheet 7: Location Intelligence
    loc_intel = result.get("location_intelligence", {})
    li_h = ["Location", "Best Title", "Title Avg NR", "Best Category", "Cat Avg NR",
            "Best Day", "Day Avg NR", "Best Combo", "Combo Avg NR", "Multiplier", "Mult Source"]
    li_rows: list[list[str]] = [li_h]
    for _, info in sorted(loc_intel.items()):
        li_rows.append([_safe_cell(info.get("Location", "")),
            _safe_cell(info.get("best_title", "")),
            _safe_cell(info.get("best_title_avg_nr", info.get("title_avg_nr", 0))),
            _safe_cell(info.get("best_category", "")),
            _safe_cell(info.get("best_category_avg_nr", info.get("cat_avg_nr", 0))),
            _safe_cell(info.get("best_day", "")),
            _safe_cell(info.get("best_day_avg_nr", info.get("day_avg_nr", 0))),
            _safe_cell(info.get("best_combo", "")),
            _safe_cell(info.get("best_combo_avg_nr", info.get("combo_avg_nr", 0))),
            _safe_cell(info.get("multiplier", "")), _safe_cell(info.get("mult_source", ""))])
    sheets["Location Intelligence"] = li_rows
    # Sheet 8: Frequency Optimisation
    fo_h = ["Combo", "Optimal/Week", "Expected Weekly NR", "NR at 1x",
            "Extra NR", "Max Observed", "NR Curve"]
    fo_rows: list[list[str]] = [fo_h]
    for r in result.get("frequency_optimization", []):
        fo_rows.append([_safe_cell(r.get("combo", "")),
            _safe_cell(r.get("optimal_posts_per_week", "")),
            _safe_cell(r.get("expected_weekly_nr", 0)), _safe_cell(r.get("nr_at_1x", 0)),
            _safe_cell(r.get("extra_nr_vs_1x", 0)),
            _safe_cell(r.get("max_observed_posts_wk", "")), _safe_cell(r.get("nr_curve", ""))])
    sheets["Frequency Optimisation"] = fo_rows
    # Sheet 9: All Runs
    ar_h = ["Post ID", "Location", "Title", "Category", "D1 Date", "Last Date",
            "Day Posted", "Run Length", "D1 Cost", "D1 Applies", "D1 NR",
            "Total Applies", "Total NR", "Profit %", "Impr Drop %",
            "Est Lifetime NR", "Multiplier", "Mult Source", "Decision", "Trigger"]
    ar_k = ["Post ID", "Location", "Title", "Category", "D1_Date", "Last_Date",
            "DayOfWeek_Posted", "Run_Length", "D1_Cost", "D1_Applies", "D1_NR",
            "Total_Applies", "Total_NR", "Profit_Pct", "Impr_Drop_Pct",
            "Est_Lifetime_NR", "Multiplier_Used", "Mult_Source", "Decision", "Trigger_Reason"]
    ar_rows: list[list[str]] = [ar_h]
    for r in result.get("all_runs", []):
        row: list[str] = []
        for k in ar_k:
            v = r.get(k)
            if k in ("Profit_Pct", "Impr_Drop_Pct") and isinstance(v, (int, float)):
                row.append(f"{v:.1f}%")
            else:
                row.append(_safe_cell(v))
        ar_rows.append(row)
    sheets["All Runs"] = ar_rows
    return sheets


def _create_google_sheet(job_data: dict[str, Any]) -> str:
    """Create a Google Spreadsheet with 9 sheets matching the Excel download.

    Args:
        job_data: The full analysis result dict from job_store.

    Returns:
        URL of the created spreadsheet.

    Raises:
        RuntimeError: If credentials are missing or API calls fail.
    """
    token = _get_gsheets_access_token()
    if not token:
        raise RuntimeError("Google Sheets not configured or token exchange failed")
    sheet_data = _job_result_to_sheet_data(job_data)
    jid = job_data.get("job_id", "unknown")
    specs: list[dict[str, Any]] = [
        {"properties": {"sheetId": idx, "title": name, "index": idx}}
        for idx, name in enumerate(sheet_data.keys())
    ]
    title = f"CG Automation Report ({jid[:8]}) - {datetime.utcnow().strftime('%Y-%m-%d')}"
    api_res = _gsheets_request(
        "POST", _GSHEETS_BASE,
        body={"properties": {"title": title}, "sheets": specs}, token=token)
    if not api_res:
        raise RuntimeError("Failed to create Google Spreadsheet")
    sid: str = api_res.get("spreadsheetId") or ""
    if not sid:
        raise RuntimeError("Google Sheets API returned no spreadsheetId")
    vr: list[dict[str, Any]] = [
        {"range": f"'{sn}'!A1", "majorDimension": "ROWS", "values": rows}
        for sn, rows in sheet_data.items() if rows
    ]
    if vr:
        br = _gsheets_request(
            "POST", f"{_GSHEETS_BASE}/{sid}/values:batchUpdate",
            body={"valueInputOption": "USER_ENTERED", "data": vr}, token=token)
        if not br:
            logger.warning("Spreadsheet created but data population failed")
    fmt: list[dict[str, Any]] = []
    for idx, (_sn, rows) in enumerate(sheet_data.items()):
        if not rows:
            continue
        fmt.append({"updateSheetProperties": {
            "properties": {"sheetId": idx, "gridProperties": {"frozenRowCount": 1}},
            "fields": "gridProperties.frozenRowCount"}})
        fmt.append({"repeatCell": {
            "range": {"sheetId": idx, "startRowIndex": 0, "endRowIndex": 1},
            "cell": {"userEnteredFormat": {
                "textFormat": {"bold": True,
                               "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                "backgroundColor": {"red": 0.059, "green": 0.125, "blue": 0.216},
                "horizontalAlignment": "CENTER"}},
            "fields": "userEnteredFormat(textFormat,backgroundColor,horizontalAlignment)"}})
        fmt.append({"autoResizeDimensions": {"dimensions": {
            "sheetId": idx, "dimension": "COLUMNS",
            "startIndex": 0, "endIndex": len(rows[0]) if rows else 10}}})
    if fmt:
        _gsheets_request("POST", f"{_GSHEETS_BASE}/{sid}:batchUpdate",
                         body={"requests": fmt}, token=token)
    _gsheets_request("POST", f"{_GDRIVE_BASE}/{sid}/permissions",
                     body={"role": "reader", "type": "anyone"}, token=token)
    return f"https://docs.google.com/spreadsheets/d/{sid}"


# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------


class InsightRequest(BaseModel):
    """Payload for the /api/insights endpoint."""

    location: str
    recommended_title: str
    title_avg_nr: float
    recommended_category: str
    cat_avg_nr: float
    best_day: str
    best_day_nr: float
    today_day: str
    profit_pct: float
    tier: int
    impr_drop_pct: float
    trigger_reason: str
    est_lifetime_nr: float
    multiplier: float
    mult_source: str
    mult_runs: int
    optimal_posts_per_week: int


class ScheduleRequest(BaseModel):
    """Payload for the POST /api/schedule endpoint."""

    job_id: str
    cron_expression: str = "0 6 * * 1"
    webhook_url: Optional[str] = None


# ---------------------------------------------------------------------------
# Constants -- Scheduler
# ---------------------------------------------------------------------------
MAX_SCHEDULED_JOBS: int = 10

# Day-of-week mapping (cron 0=Sunday convention)
_DOW_MAP: dict[int, int] = {
    0: 6,  # cron Sunday -> Python weekday 6
    1: 0,  # cron Monday -> Python weekday 0
    2: 1,
    3: 2,
    4: 3,
    5: 4,
    6: 5,
    7: 6,  # cron 7 also = Sunday
}

# ---------------------------------------------------------------------------
# In-memory schedule store + helpers
# ---------------------------------------------------------------------------
_schedule_lock = threading.Lock()
_schedule_store: dict[str, dict[str, Any]] = {}


def _parse_simple_cron(expr: str) -> int:
    """Parse a simplified cron expression and return interval in seconds.

    Supports the format ``0 H * * D`` where *H* is hour (0-23) and *D* is
    day-of-week (0-7, 0 and 7 = Sunday).  The interval is always exactly
    7 days (604800 seconds) because only one weekday is specified.

    Args:
        expr: Cron-style string, e.g. ``"0 6 * * 1"``.

    Returns:
        Interval in seconds (always 604800 for a weekly schedule).

    Raises:
        ValueError: If the expression does not match the supported format.
    """
    parts = expr.strip().split()
    if len(parts) != 5:
        raise ValueError(
            f"Cron expression must have 5 fields, got {len(parts)}: '{expr}'"
        )
    minute, hour, dom, month, dow = parts

    if minute != "0":
        raise ValueError("Only minute=0 is supported in simplified cron")
    if dom != "*" or month != "*":
        raise ValueError("Only day-of-month=* and month=* are supported")

    try:
        hour_int = int(hour)
    except ValueError as exc:
        raise ValueError(f"Invalid hour '{hour}' in cron expression") from exc
    if not 0 <= hour_int <= 23:
        raise ValueError(f"Hour must be 0-23, got {hour_int}")

    try:
        dow_int = int(dow)
    except ValueError as exc:
        raise ValueError(f"Invalid day-of-week '{dow}' in cron expression") from exc
    if dow_int not in _DOW_MAP:
        raise ValueError(f"Day-of-week must be 0-7, got {dow_int}")

    # Weekly interval
    return 7 * 24 * 60 * 60


def _seconds_until_next(cron_expr: str) -> float:
    """Return seconds from now until the next occurrence of the cron time.

    Args:
        cron_expr: Simplified cron string ``"0 H * * D"``.

    Returns:
        Positive float representing seconds until the next matching moment.
    """
    parts = cron_expr.strip().split()
    hour_int = int(parts[1])
    dow_int = int(parts[4])
    target_weekday = _DOW_MAP[dow_int]

    now = datetime.now()
    days_ahead = target_weekday - now.weekday()
    if days_ahead < 0:
        days_ahead += 7

    candidate = now.replace(
        hour=hour_int, minute=0, second=0, microsecond=0
    ) + timedelta(days=days_ahead)

    if candidate <= now:
        candidate += timedelta(weeks=1)

    return (candidate - now).total_seconds()


def _fire_scheduled_job(schedule_id: str) -> None:
    """Execute a scheduled re-analysis and optionally POST results to webhook.

    Runs inside a daemon thread spawned by ``threading.Timer``.  Re-runs
    ``engine.run_analysis`` on the stored source DataFrame, updates the
    job store with fresh results, and reschedules itself.

    Args:
        schedule_id: Key into ``_schedule_store``.
    """
    with _schedule_lock:
        entry = _schedule_store.get(schedule_id)
        if entry is None:
            return  # cancelled

    job_id: str = entry["job_id"]
    webhook_url: str | None = entry.get("webhook_url")
    cron_expr: str = entry["cron_expression"]

    logger.info("Scheduled job %s firing for job_id=%s", schedule_id, job_id)

    stored = job_store.get(job_id)
    if stored is None:
        logger.warning(
            "Scheduled job %s: original job_id=%s no longer in store; skipping",
            schedule_id, job_id,
        )
    else:
        last_df = stored.get("_source_df")
        sell_cpa: float = stored.get("_sell_cpa", 1.20)
        if last_df is not None:
            try:
                result = engine.run_analysis(last_df.copy(), sell_cpa=sell_cpa)
                result = _sanitize_for_json(result)
                result["job_id"] = job_id
                result["_source_df"] = last_df
                result["_sell_cpa"] = sell_cpa
                job_store[job_id] = result
                logger.info("Scheduled re-analysis complete for job_id=%s", job_id)

                if webhook_url:
                    scorecard = result.get("scorecard", {})
                    payload = {
                        "schedule_id": schedule_id,
                        "job_id": job_id,
                        "ran_at": datetime.now().isoformat(),
                        "scorecard": scorecard,
                    }
                    try:
                        data = json.dumps(payload).encode("utf-8")
                        wh_req = urllib.request.Request(
                            webhook_url,
                            data=data,
                            headers={"Content-Type": "application/json"},
                            method="POST",
                        )
                        with urllib.request.urlopen(wh_req, timeout=15) as wh_resp:
                            logger.info(
                                "Webhook POST to %s returned status %d",
                                webhook_url, wh_resp.status,
                            )
                    except Exception:
                        logger.error(
                            "Webhook POST to %s failed", webhook_url, exc_info=True,
                        )
            except Exception:
                logger.error(
                    "Scheduled re-analysis failed for job_id=%s", job_id, exc_info=True,
                )
        else:
            logger.warning(
                "Scheduled job %s: no source DataFrame for job_id=%s",
                schedule_id, job_id,
            )

    # Reschedule for next week
    with _schedule_lock:
        if schedule_id not in _schedule_store:
            return  # cancelled while running

        delay = _seconds_until_next(cron_expr)
        timer = threading.Timer(delay, _fire_scheduled_job, args=[schedule_id])
        timer.daemon = True
        timer.start()
        _schedule_store[schedule_id]["timer"] = timer
        _schedule_store[schedule_id]["next_run"] = (
            datetime.now() + timedelta(seconds=delay)
        ).isoformat()
        _schedule_store[schedule_id]["last_run"] = datetime.now().isoformat()
        logger.info("Rescheduled %s: next run in %.0f seconds", schedule_id, delay)


# ===================================================================
# EXCEL GENERATION -- 9 sheets (Section 19)
# ===================================================================


def _style_header(ws: Any, col_count: int) -> None:
    """Apply dark navy header styling to row 1."""
    header_fill = PatternFill(
        start_color=CLR_HEADER_BG, end_color=CLR_HEADER_BG, fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )


def _apply_tier_row_color(
    ws: Any, row_idx: int, tier: int, col_count: int
) -> None:
    """Apply tier-specific background colour to a row."""
    bg = TIER_BG_MAP.get(tier)
    if bg:
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        for c in range(1, col_count + 1):
            ws.cell(row=row_idx, column=c).fill = fill


def _apply_nr_font(cell: Any, value: float | None) -> None:
    """Colour NR values green (positive) or red (negative)."""
    if value is None:
        return
    if value > 0:
        cell.font = Font(color=CLR_POS_NR)
    elif value < 0:
        cell.font = Font(color=CLR_NEG_NR)


def _set_column_widths(ws: Any, widths: list[int]) -> None:
    """Set column widths."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_sheet_from_records(
    wb: Workbook,
    sheet_name: str,
    headers: list[str],
    records: list[dict],
    key_map: list[str],
    col_widths: list[int] | None = None,
    nr_columns: list[int] | None = None,
    tier_column: int | None = None,
    pct_columns: list[int] | None = None,
    money_columns: list[int] | None = None,
) -> None:
    """Write a sheet from a list of dicts with styling.

    Args:
        wb: Target workbook.
        sheet_name: Tab name.
        headers: Column header labels.
        records: List of dicts to write.
        key_map: Dict keys matching each header column.
        col_widths: Optional per-column widths.
        nr_columns: 1-based column indices to format as NR ($).
        tier_column: 1-based column index containing tier for row colouring.
        pct_columns: 1-based column indices to format as percentages.
        money_columns: 1-based column indices to format as currency.
    """
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # Headers
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, len(headers))

    # Data rows
    for ri, rec in enumerate(records, 2):
        for ci, key in enumerate(key_map, 1):
            val = rec.get(key)
            cell = ws.cell(row=ri, column=ci, value=val)

            if money_columns and ci in money_columns and isinstance(val, (int, float)):
                cell.number_format = "$#,##0.00"
                _apply_nr_font(cell, val)
            elif pct_columns and ci in pct_columns and isinstance(val, (int, float)):
                cell.number_format = "0.0%"
                cell.value = val / 100.0 if val is not None else None
            elif nr_columns and ci in nr_columns and isinstance(val, (int, float)):
                cell.number_format = "$#,##0.00"
                _apply_nr_font(cell, val)

        # Tier row colouring
        if tier_column is not None:
            tier_val = rec.get(key_map[tier_column - 1])
            if isinstance(tier_val, (int, float)) and tier_val in TIER_BG_MAP:
                _apply_tier_row_color(ws, ri, int(tier_val), len(headers))

    if col_widths:
        _set_column_widths(ws, col_widths)


def _build_conflict_counts(result: dict[str, Any]) -> dict[str, int]:
    """Count location conflicts for the daily action plan sheet."""
    counts: dict[str, int] = {}
    for c in result.get("location_conflicts", []):
        loc = c.get("Location", c.get("location", ""))
        lk = str(loc).strip().lower()
        counts[lk] = counts.get(lk, 0) + 1
    return counts


def generate_excel(result: dict[str, Any]) -> io.BytesIO:
    """Generate the full 9-sheet styled Excel workbook.

    Args:
        result: The full analysis result dict from engine.run_analysis().

    Returns:
        BytesIO buffer containing the .xlsx file.
    """
    wb = Workbook()
    wb.remove(wb.active)

    conflict_counts = _build_conflict_counts(result)

    # ---- Sheet 1: Daily Action Plan ----
    dap_headers = [
        "#", "Location", "Best Title", "Best Category", "Best Day",
        "Today Good?", "Tier", "Trigger", "Est D1 NR", "Est Lifetime NR",
        "Posts/Week", "Multiplier", "Mult Source", "Competing Filtered",
        "Last Run Date",
    ]
    ws1 = wb.create_sheet(title="Daily Action Plan")
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A2"
    for ci, h in enumerate(dap_headers, 1):
        ws1.cell(row=1, column=ci, value=h)
    _style_header(ws1, len(dap_headers))

    dap = result.get("daily_action_plan", [])
    for ri, rec in enumerate(dap, 2):
        num = ri - 1
        loc = rec.get("location", rec.get("Location", ""))
        ws1.cell(row=ri, column=1, value=num)
        ws1.cell(row=ri, column=2, value=loc)
        ws1.cell(row=ri, column=3, value=rec.get("recommended_title", rec.get("Best_Title", "")))
        ws1.cell(row=ri, column=4, value=rec.get("recommended_category", rec.get("Best_Category", "")))
        ws1.cell(row=ri, column=5, value=rec.get("best_day", rec.get("Best_Day", "")))
        ws1.cell(row=ri, column=6, value=rec.get("today_good", rec.get("Today_Good", "")))
        tier = rec.get("tier", rec.get("Tier", 0))
        ws1.cell(row=ri, column=7, value=tier)
        ws1.cell(row=ri, column=8, value=rec.get("trigger_reason", rec.get("Trigger_Reason", "")))

        d1_nr = rec.get("d1_nr", rec.get("D1_NR", rec.get("Est_D1_NR", 0)))
        d1_cell = ws1.cell(row=ri, column=9, value=d1_nr)
        d1_cell.number_format = "$#,##0.00"
        _apply_nr_font(d1_cell, d1_nr)

        life_nr = rec.get("est_lifetime_nr", rec.get("Est_Lifetime_NR", 0))
        life_cell = ws1.cell(row=ri, column=10, value=life_nr)
        life_cell.number_format = "$#,##0.00"
        _apply_nr_font(life_cell, life_nr)

        ws1.cell(row=ri, column=11, value=rec.get("optimal_posts_per_week", rec.get("Optimal_Posts_Per_Week", 1)))
        ws1.cell(row=ri, column=12, value=rec.get("multiplier", rec.get("Multiplier_Used", "")))
        ws1.cell(row=ri, column=13, value=rec.get("mult_source", rec.get("Mult_Source", "")))

        lk = str(loc).strip().lower()
        ws1.cell(row=ri, column=14, value=conflict_counts.get(lk, 0))
        ws1.cell(row=ri, column=15, value=rec.get("last_run_date", rec.get("D1_Date", "")))

        if isinstance(tier, (int, float)) and int(tier) in TIER_BG_MAP:
            _apply_tier_row_color(ws1, ri, int(tier), len(dap_headers))

    _set_column_widths(ws1, [5, 20, 35, 18, 12, 12, 6, 40, 12, 14, 10, 10, 14, 12, 12])

    # ---- Sheet 2: All Repost Candidates ----
    rp_headers = [
        "#", "Location", "Title", "Category", "Tier", "Cost", "Profit %",
        "Trigger", "D1 Applies", "D1 NR", "Lifetime NR", "Multiplier",
        "Mult Source", "Best Day", "Today Good?", "Posts/Week",
    ]
    ws2 = wb.create_sheet(title="All Repost Candidates")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A2"
    for ci, h in enumerate(rp_headers, 1):
        ws2.cell(row=1, column=ci, value=h)
    _style_header(ws2, len(rp_headers))

    for ri, rec in enumerate(result.get("all_repost", []), 2):
        ws2.cell(row=ri, column=1, value=ri - 1)
        ws2.cell(row=ri, column=2, value=rec.get("Location", ""))
        ws2.cell(row=ri, column=3, value=rec.get("Title", ""))
        ws2.cell(row=ri, column=4, value=rec.get("Category", ""))
        tier = rec.get("Tier", 0)
        ws2.cell(row=ri, column=5, value=tier)

        cost = rec.get("D1_Cost", 0)
        cost_cell = ws2.cell(row=ri, column=6, value=cost)
        cost_cell.number_format = "$#,##0.00"

        pct = rec.get("Profit_Pct")
        pct_cell = ws2.cell(row=ri, column=7, value=(pct / 100.0 if pct is not None else None))
        pct_cell.number_format = "0.0%"

        ws2.cell(row=ri, column=8, value=rec.get("Trigger_Reason", ""))
        ws2.cell(row=ri, column=9, value=rec.get("D1_Applies", 0))

        d1nr = rec.get("D1_NR", 0)
        d1nr_cell = ws2.cell(row=ri, column=10, value=d1nr)
        d1nr_cell.number_format = "$#,##0.00"
        _apply_nr_font(d1nr_cell, d1nr)

        life_nr = rec.get("Est_Lifetime_NR", 0)
        life_cell = ws2.cell(row=ri, column=11, value=life_nr)
        life_cell.number_format = "$#,##0.00"
        _apply_nr_font(life_cell, life_nr)

        ws2.cell(row=ri, column=12, value=rec.get("Multiplier_Used", ""))
        ws2.cell(row=ri, column=13, value=rec.get("Mult_Source", ""))
        ws2.cell(row=ri, column=14, value=rec.get("Best_Day", ""))
        ws2.cell(row=ri, column=15, value=rec.get("Today_Good", ""))
        ws2.cell(row=ri, column=16, value=rec.get("Optimal_Posts_Per_Week", 1))

        if isinstance(tier, (int, float)) and int(tier) in TIER_BG_MAP:
            _apply_tier_row_color(ws2, ri, int(tier), len(rp_headers))

    _set_column_widths(ws2, [5, 20, 35, 18, 6, 10, 10, 40, 10, 12, 14, 10, 14, 12, 12, 10])

    # ---- Sheet 3: Best Per Location ----
    _write_sheet_from_records(
        wb,
        "Best Per Location",
        headers=["#", "Location", "Best Title", "Best Category", "Tier",
                 "Est Lifetime NR", "Best Day", "Posts/Week"],
        records=[
            {**r, "_num": i}
            for i, r in enumerate(result.get("best_per_location", []), 1)
        ],
        key_map=["_num", "Location", "Title", "Category", "Tier",
                 "Est_Lifetime_NR", "DayOfWeek_Posted", "Run_Length"],
        col_widths=[5, 20, 35, 18, 6, 14, 12, 10],
        nr_columns=[6],
        tier_column=5,
    )

    # ---- Sheet 4: Location Conflicts ----
    _write_sheet_from_records(
        wb,
        "Location Conflicts",
        headers=["Location", "Title", "Category", "Tier", "Est Lifetime NR",
                 "Lost To", "NR Gap"],
        records=result.get("location_conflicts", []),
        key_map=["Location", "Title", "Category", "Tier", "Est_Lifetime_NR",
                 "lost_to", "nr_gap"],
        col_widths=[20, 35, 18, 6, 14, 40, 10],
        nr_columns=[5, 7],
        tier_column=4,
    )

    # ---- Sheet 5: Keep Running ----
    _write_sheet_from_records(
        wb,
        "Keep Running",
        headers=["Location", "Title", "Category", "Total NR", "Profit %",
                 "Trigger"],
        records=result.get("keep_running", []),
        key_map=["Location", "Title", "Category", "Total_NR", "Profit_Pct",
                 "Trigger_Reason"],
        col_widths=[20, 35, 18, 12, 10, 40],
        nr_columns=[4],
        pct_columns=[5],
    )

    # ---- Sheet 6: Skip ----
    _write_sheet_from_records(
        wb,
        "Skip",
        headers=["Location", "Title", "Category", "Total NR", "Profit %",
                 "Skip Reason"],
        records=result.get("skip", []),
        key_map=["Location", "Title", "Category", "Total_NR", "Profit_Pct",
                 "Trigger_Reason"],
        col_widths=[20, 35, 18, 12, 10, 50],
        nr_columns=[4],
        pct_columns=[5],
    )

    # ---- Sheet 7: Location Intelligence ----
    loc_intel = result.get("location_intelligence", {})
    ws7 = wb.create_sheet(title="Location Intelligence")
    ws7.sheet_view.showGridLines = False
    ws7.freeze_panes = "A2"
    li_headers = [
        "Location", "Best Title", "Title Avg NR", "Best Category",
        "Cat Avg NR", "Best Day", "Day Avg NR", "Best Combo",
        "Combo Avg NR", "Multiplier", "Mult Source",
    ]
    for ci, h in enumerate(li_headers, 1):
        ws7.cell(row=1, column=ci, value=h)
    _style_header(ws7, len(li_headers))

    for ri, (_, info) in enumerate(sorted(loc_intel.items()), 2):
        ws7.cell(row=ri, column=1, value=info.get("Location", ""))
        ws7.cell(row=ri, column=2, value=info.get("best_title", ""))

        title_nr = info.get("best_title_avg_nr", info.get("title_avg_nr", 0))
        c3 = ws7.cell(row=ri, column=3, value=title_nr)
        c3.number_format = "$#,##0.00"
        _apply_nr_font(c3, title_nr)

        ws7.cell(row=ri, column=4, value=info.get("best_category", ""))

        cat_nr = info.get("best_category_avg_nr", info.get("cat_avg_nr", 0))
        c5 = ws7.cell(row=ri, column=5, value=cat_nr)
        c5.number_format = "$#,##0.00"
        _apply_nr_font(c5, cat_nr)

        ws7.cell(row=ri, column=6, value=info.get("best_day", ""))

        day_nr = info.get("best_day_avg_nr", info.get("day_avg_nr", 0))
        c7 = ws7.cell(row=ri, column=7, value=day_nr)
        c7.number_format = "$#,##0.00"
        _apply_nr_font(c7, day_nr)

        ws7.cell(row=ri, column=8, value=info.get("best_combo", ""))

        combo_nr = info.get("best_combo_avg_nr", info.get("combo_avg_nr", 0))
        c9 = ws7.cell(row=ri, column=9, value=combo_nr)
        c9.number_format = "$#,##0.00"
        _apply_nr_font(c9, combo_nr)

        ws7.cell(row=ri, column=10, value=info.get("multiplier", ""))

        mult_src = info.get("mult_source", "")
        c11 = ws7.cell(row=ri, column=11, value=mult_src)
        if mult_src == "Location Avg":
            c11.fill = PatternFill(
                start_color=CLR_LOC_MULT, end_color=CLR_LOC_MULT, fill_type="solid"
            )
        elif mult_src == "Category Avg":
            c11.fill = PatternFill(
                start_color=CLR_CAT_MULT, end_color=CLR_CAT_MULT, fill_type="solid"
            )

    _set_column_widths(ws7, [20, 35, 12, 18, 12, 12, 12, 40, 12, 10, 14])

    # ---- Sheet 8: Frequency Optimisation ----
    _write_sheet_from_records(
        wb,
        "Frequency Optimisation",
        headers=["Combo", "Optimal/Week", "Expected Weekly NR", "NR at 1x",
                 "Extra NR", "Max Observed", "NR Curve"],
        records=result.get("frequency_optimization", []),
        key_map=["combo", "optimal_posts_per_week", "expected_weekly_nr",
                 "nr_at_1x", "extra_nr_vs_1x", "max_observed_posts_wk",
                 "nr_curve"],
        col_widths=[50, 12, 16, 12, 12, 14, 50],
        nr_columns=[3, 4, 5],
    )

    # ---- Sheet 9: All Runs ----
    ar_headers = [
        "Post ID", "Location", "Title", "Category", "D1 Date", "Last Date",
        "Day Posted", "Run Length", "D1 Cost", "D1 Applies", "D1 NR",
        "Total Applies", "Total NR", "Profit %", "Impr Drop %",
        "Est Lifetime NR", "Multiplier", "Mult Source", "Decision", "Trigger",
    ]
    ar_keys = [
        "Post ID", "Location", "Title", "Category", "D1_Date", "Last_Date",
        "DayOfWeek_Posted", "Run_Length", "D1_Cost", "D1_Applies", "D1_NR",
        "Total_Applies", "Total_NR", "Profit_Pct", "Impr_Drop_Pct",
        "Est_Lifetime_NR", "Multiplier_Used", "Mult_Source", "Decision",
        "Trigger_Reason",
    ]
    _write_sheet_from_records(
        wb,
        "All Runs",
        headers=ar_headers,
        records=result.get("all_runs", []),
        key_map=ar_keys,
        col_widths=[
            12, 20, 35, 18, 12, 12, 12, 10, 10, 10, 12, 10, 12, 10, 10,
            14, 10, 14, 12, 40,
        ],
        nr_columns=[11, 13, 16],
        money_columns=[9],
        pct_columns=[14, 15],
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ===================================================================
# JSON SANITIZATION (numpy/pandas types -> native Python)
# ===================================================================


def _sanitize_for_json(obj: Any) -> Any:
    """Recursively convert numpy/pandas types to native Python for JSON."""
    if isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize_for_json(v) for v in obj]
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        val = float(obj)
        if math.isnan(val) or math.isinf(val):
            return None
        return val
    if isinstance(obj, np.bool_):
        return bool(obj)
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    if isinstance(obj, pd.Timestamp):
        return obj.isoformat()
    if hasattr(obj, 'isoformat'):
        return obj.isoformat()
    return obj


# ===================================================================
# API ENDPOINTS
# ===================================================================


@app.get("/api/health")
async def health() -> dict[str, str]:
    """Health check endpoint."""
    return {"status": "ok", "version": "1.0.0"}


def _persist_to_supabase(
    result: dict[str, Any], job_id: str, filename: str, sell_cpa: float
) -> None:
    """Persist job + action plan to Supabase in a background thread.

    Non-blocking: failures are logged but never raise.

    Args:
        result: Sanitised analysis result dict.
        job_id: UUID of the stored job.
        filename: Original uploaded filename.
        sell_cpa: Revenue per apply used.
    """
    def _do_persist() -> None:
        try:
            scorecard: dict[str, Any] = result.get("scorecard", {})
            supabase_store.save_job(job_id, filename, sell_cpa, scorecard)

            action_plan: list[dict[str, Any]] = result.get("daily_action_plan", [])
            if action_plan:
                supabase_store.save_action_plan(job_id, action_plan)

            logger.info("Supabase persistence complete for job_id=%s", job_id)
        except Exception:
            logger.error("Supabase persistence failed for job_id=%s", job_id, exc_info=True)

    thread = threading.Thread(target=_do_persist, daemon=True)
    thread.start()


def _safe_call(fn: Any, *args: Any, **kwargs: Any) -> None:
    """Call *fn* swallowing all exceptions (for background threads).

    Args:
        fn: Callable to invoke.
        *args: Positional arguments.
        **kwargs: Keyword arguments.
    """
    try:
        fn(*args, **kwargs)
    except Exception:
        logger.error("Background Supabase call %s failed", fn.__name__, exc_info=True)


@app.post("/api/analyse")
async def api_analyse(
    file: UploadFile = File(...),
    sell_cpa: float = 1.20,
) -> dict[str, Any]:
    """Upload an Excel file and run the full CG Automation analysis.

    Args:
        file: Uploaded Excel file (multipart/form-data).
        sell_cpa: Revenue per apply in USD (varies by client/campaign, default $1.20).

    Returns:
        Full JSON analysis with scorecard, daily action plan, and all views.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file uploaded")

    try:
        contents = await file.read()
        fname = (file.filename or "").lower()
        if fname.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents))
    except Exception as exc:
        logger.error("Failed to read uploaded file", exc_info=True)
        raise HTTPException(
            status_code=400, detail=f"Failed to read file: {exc}"
        ) from exc

    try:
        result = engine.run_analysis(df, sell_cpa=sell_cpa)
    except ValueError as exc:
        logger.error("Validation error during analysis", exc_info=True)
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.error("Analysis failed", exc_info=True)
        raise HTTPException(
            status_code=500, detail=f"Analysis error: {exc}"
        ) from exc

    # Convert numpy types to native Python for JSON serialization
    result = _sanitize_for_json(result)

    job_id = str(uuid.uuid4())
    result["job_id"] = job_id
    # Store source data for scheduled re-analysis
    result["_source_df"] = df
    result["_sell_cpa"] = sell_cpa
    job_store[job_id] = result

    # Fire-and-forget Slack notification (non-blocking)
    _notify_slack_analysis(result, job_id)

    # Fire-and-forget Supabase persistence (non-blocking)
    _persist_to_supabase(result, job_id, file.filename or "", sell_cpa)

    # Return JSON-safe copy (strip internal DataFrame / sell_cpa keys)
    response = {k: v for k, v in result.items() if not k.startswith("_")}
    return response


@app.post("/api/test-slack")
async def api_test_slack() -> dict[str, str]:
    """Send a test message to the configured Slack webhook.

    Returns 200 with status even if SLACK_WEBHOOK_URL is not set (returns a warning).
    """
    if not SLACK_WEBHOOK_URL:
        return {"status": "skipped", "reason": "SLACK_WEBHOOK_URL not configured"}

    payload: dict[str, str] = {
        "text": "CG Automation: Slack integration test -- webhook is working!"
    }
    try:
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            SLACK_WEBHOOK_URL,
            data=data,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            if resp.status == 200:
                return {"status": "ok", "message": "Test message sent to Slack"}
            return {"status": "error", "message": f"Slack returned status {resp.status}"}
    except Exception as exc:
        logger.warning("Slack test failed", exc_info=True)
        return {"status": "error", "message": str(exc)}


@app.post("/api/insights")
async def api_insights(req: InsightRequest) -> dict[str, str]:
    """Generate an AI insight for a single repost candidate using Claude.

    Calls Claude claude-sonnet-4-20250514 with all per-location context fields to
    produce a concise, actionable recommendation.

    Args:
        req: InsightRequest with all per-location context fields.

    Returns:
        Dict with a single 'insight' key containing Claude's response.
    """
    user_prompt = (
        f"Location: {req.location}\n"
        f"Recommended Title: {req.recommended_title} (avg NR: ${req.title_avg_nr:.2f})\n"
        f"Recommended Category: {req.recommended_category} (avg NR: ${req.cat_avg_nr:.2f})\n"
        f"Best Day to Post here: {req.best_day} (avg NR: ${req.best_day_nr:.2f})\n"
        f"Today is: {req.today_day}\n"
        f"Historical Profit: {req.profit_pct:.1f}%\n"
        f"Tier: {req.tier}\n"
        f"Impression Drop: {req.impr_drop_pct:.1f}%\n"
        f"Repost Trigger: {req.trigger_reason}\n"
        f"Est Lifetime NR: ${req.est_lifetime_nr:.2f}\n"
        f"Location Multiplier: {req.multiplier:.2f}x ({req.mult_source}, {req.mult_runs} runs)\n"
        f"Optimal Posts This Week: {req.optimal_posts_per_week}x\n"
        f"In 2 sentences: why repost now and one specific action tip."
    )

    system_prompt = (
        "You are a Craigslist ad campaign analyst. You give specific, actionable "
        "advice based on data. Be concise. Always end with one concrete action "
        "the operator can take today."
    )

    try:
        from llm_router import generate_insight
        insight_text = generate_insight(system_prompt, user_prompt)
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Unexpected error generating insight", exc_info=True)
        raise HTTPException(
            status_code=500, detail=f"Insight generation failed: {exc}"
        ) from exc

    return {"insight": insight_text}


@app.get("/api/download/{job_id}")
async def api_download(job_id: str) -> StreamingResponse:
    """Download the styled Excel report for a completed analysis job.

    Retrieves the stored analysis result by job_id, generates a
    9-sheet styled Excel workbook, and returns it as a file download.

    Args:
        job_id: UUID of the analysis job.

    Returns:
        Streaming Excel file download with Content-Disposition header.
    """
    result = job_store.get(job_id)
    if result is None:
        raise HTTPException(
            status_code=404, detail=f"Job {job_id} not found or expired"
        )

    try:
        buf = generate_excel(result)
    except Exception as exc:
        logger.error("Excel generation failed", exc_info=True)
        raise HTTPException(
            status_code=500, detail=f"Excel generation error: {exc}"
        ) from exc

    filename = f"CG_Automation_Report_{job_id[:8]}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )


# ===================================================================
# SCHEDULED RE-ANALYSIS ENDPOINTS
# ===================================================================


@app.post("/api/schedule")
async def api_schedule(req: ScheduleRequest) -> dict[str, Any]:
    """Schedule a recurring re-analysis for an existing job.

    Creates a ``threading.Timer``-based weekly schedule that re-runs
    ``engine.run_analysis`` on the original uploaded data and optionally
    POSTs the scorecard summary to a webhook URL.

    Args:
        req: ScheduleRequest with job_id, optional cron_expression and webhook_url.

    Returns:
        Dict with schedule_id, next_run ISO timestamp, and interval_seconds.
    """
    # Validate job exists
    stored = job_store.get(req.job_id)
    if stored is None:
        raise HTTPException(
            status_code=404,
            detail=f"Job {req.job_id} not found or expired",
        )

    if stored.get("_source_df") is None:
        raise HTTPException(
            status_code=400,
            detail=f"Job {req.job_id} has no stored source data for re-analysis",
        )

    # Validate cron expression
    try:
        interval_seconds = _parse_simple_cron(req.cron_expression)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    # Enforce max schedules
    with _schedule_lock:
        if len(_schedule_store) >= MAX_SCHEDULED_JOBS:
            raise HTTPException(
                status_code=409,
                detail=f"Maximum of {MAX_SCHEDULED_JOBS} scheduled jobs reached",
            )

        schedule_id = str(uuid.uuid4())
        delay = _seconds_until_next(req.cron_expression)
        next_run = (datetime.now() + timedelta(seconds=delay)).isoformat()

        timer = threading.Timer(delay, _fire_scheduled_job, args=[schedule_id])
        timer.daemon = True
        timer.start()

        _schedule_store[schedule_id] = {
            "schedule_id": schedule_id,
            "job_id": req.job_id,
            "cron_expression": req.cron_expression,
            "webhook_url": req.webhook_url,
            "interval_seconds": interval_seconds,
            "next_run": next_run,
            "last_run": None,
            "created_at": datetime.now().isoformat(),
            "timer": timer,
        }

    logger.info(
        "Created schedule %s for job %s (next run in %.0f s)",
        schedule_id, req.job_id, delay,
    )

    # Persist schedule to Supabase (background, non-blocking)
    threading.Thread(
        target=lambda: _safe_call(
            supabase_store.save_schedule,
            schedule_id, req.job_id, req.cron_expression,
            req.webhook_url or "", next_run,
        ),
        daemon=True,
    ).start()

    return {
        "schedule_id": schedule_id,
        "job_id": req.job_id,
        "cron_expression": req.cron_expression,
        "interval_seconds": interval_seconds,
        "next_run": next_run,
    }


@app.get("/api/schedules")
async def api_list_schedules() -> list[dict[str, Any]]:
    """List all active scheduled re-analysis jobs.

    Returns:
        List of schedule entries (timer object excluded from response).
    """
    with _schedule_lock:
        schedules: list[dict[str, Any]] = []
        for entry in _schedule_store.values():
            schedules.append(
                {k: v for k, v in entry.items() if k != "timer"}
            )
    logger.info("Listed %d active schedules", len(schedules))
    return schedules


@app.delete("/api/schedule/{schedule_id}")
async def api_cancel_schedule(schedule_id: str) -> dict[str, str]:
    """Cancel and remove a scheduled re-analysis job.

    Args:
        schedule_id: UUID of the schedule to cancel.

    Returns:
        Confirmation dict with schedule_id and status.
    """
    with _schedule_lock:
        entry = _schedule_store.pop(schedule_id, None)

    if entry is None:
        raise HTTPException(
            status_code=404,
            detail=f"Schedule {schedule_id} not found",
        )

    timer: threading.Timer = entry.get("timer")
    if timer is not None:
        timer.cancel()

    logger.info("Cancelled schedule %s (job_id=%s)", schedule_id, entry["job_id"])

    # Remove from Supabase (background, non-blocking)
    threading.Thread(
        target=lambda: _safe_call(supabase_store.delete_schedule, schedule_id),
        daemon=True,
    ).start()

    return {"schedule_id": schedule_id, "status": "cancelled"}


# ---------------------------------------------------------------------------
# Supabase read endpoints
# ---------------------------------------------------------------------------


@app.get("/api/jobs")
async def api_list_jobs(limit: int = 20) -> list[dict[str, Any]]:
    """List recent analysis jobs from Supabase.

    Falls back to in-memory job_store keys if Supabase is not configured.

    Args:
        limit: Maximum number of jobs to return.

    Returns:
        List of job summary dicts.
    """
    sb_jobs = supabase_store.list_jobs(limit=limit)
    if sb_jobs:
        return sb_jobs

    # Fallback: return summaries from in-memory store
    results: list[dict[str, Any]] = []
    for jid, data in list(job_store.items())[:limit]:
        sc = data.get("scorecard", {})
        results.append({
            "job_id": jid,
            "filename": data.get("filename", ""),
            "status": "completed",
            "total_locations": sc.get("total_locations", 0),
            "total_spend": sc.get("total_spend") or sc.get("total_cost", 0),
            "total_nr": sc.get("total_nr") or sc.get("total_lifetime_nr", 0),
        })
    return results


@app.get("/api/job/{job_id}")
async def api_get_job(job_id: str) -> dict[str, Any]:
    """Get a single job by job_id.

    Tries Supabase first, then falls back to in-memory store.

    Args:
        job_id: UUID of the job.

    Returns:
        Job dict with scorecard and action plan data.
    """
    sb_job = supabase_store.get_job(job_id)
    if sb_job:
        return sb_job

    # Fallback: in-memory store
    stored = job_store.get(job_id)
    if stored is None:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")

    response = {k: v for k, v in stored.items() if not k.startswith("_")}
    return response


# ===================================================================
# GOOGLE SHEETS EXPORT ENDPOINTS
# ===================================================================


@app.post("/api/sheets/{job_id}")
async def api_create_sheets(job_id: str) -> dict[str, Any]:
    """Export analysis results to a new Google Spreadsheet.

    Creates a Google Sheet with the same 9 tabs as the Excel download.
    Runs the Sheets API calls in a background thread and stores the
    resulting URL in job_store for polling via GET /api/sheets-url/{job_id}.

    If GOOGLE_SHEETS_CREDENTIALS_B64 is not set, returns 501.

    Args:
        job_id: UUID of the analysis job.

    Returns:
        Dict with sheets_url if already created, or status=pending.
    """
    if not GOOGLE_SHEETS_CREDENTIALS_B64:
        raise HTTPException(
            status_code=501,
            detail="Google Sheets not configured (GOOGLE_SHEETS_CREDENTIALS_B64 not set)",
        )

    job_data = job_store.get(job_id)
    if job_data is None:
        raise HTTPException(
            status_code=404, detail=f"Job {job_id} not found or expired"
        )

    # If already exported, return the existing URL
    existing_url = job_data.get("sheets_url")
    if existing_url:
        return {"sheets_url": existing_url, "status": "ready"}

    # Run in background thread to avoid blocking the event loop
    def _bg_create() -> None:
        try:
            url = _create_google_sheet(job_data)
            job_data["sheets_url"] = url
            logger.info("Google Sheet created for job %s: %s", job_id, url)
        except Exception as exc:
            job_data["sheets_error"] = str(exc)
            logger.error(
                "Google Sheet creation failed for job %s: %s",
                job_id, exc, exc_info=True,
            )

    bg_thread = threading.Thread(target=_bg_create, daemon=True)
    bg_thread.start()

    return {
        "status": "pending",
        "message": "Google Sheet creation started. Poll GET /api/sheets-url/{job_id}.",
    }


@app.get("/api/sheets-url/{job_id}")
async def api_get_sheets_url(job_id: str) -> dict[str, Any]:
    """Poll for the Google Sheets URL after POST /api/sheets/{job_id}.

    Args:
        job_id: UUID of the analysis job.

    Returns:
        Dict with sheets_url if ready, status=pending if still creating,
        or error details if creation failed.
    """
    if not GOOGLE_SHEETS_CREDENTIALS_B64:
        raise HTTPException(
            status_code=501,
            detail="Google Sheets not configured (GOOGLE_SHEETS_CREDENTIALS_B64 not set)",
        )

    job_data = job_store.get(job_id)
    if job_data is None:
        raise HTTPException(
            status_code=404, detail=f"Job {job_id} not found or expired"
        )

    sheets_url = job_data.get("sheets_url")
    if sheets_url:
        return {"sheets_url": sheets_url, "status": "ready"}

    sheets_error = job_data.get("sheets_error")
    if sheets_error:
        return {"status": "error", "detail": sheets_error}

    return {"status": "pending", "message": "Google Sheet is still being created."}


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import uvicorn

    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
